"""
Voucher import/export service.

Bulk-create `VoucherCode` rows from a CSV or XLSX upload, and produce
the same shape on export so files round-trip cleanly.

The shop admin uploads a file plus a set of voucher settings (discount
type / value / expiry / restrictions / etc.) on page 1, lands on page 2
to map the file's columns to a small set of mappable fields, picks a
duplicate strategy, and confirms. Every imported voucher inherits the
batch settings; the only per-row inputs are the mapped columns.

Mappable columns (per the approved plan):
  - code         required — `VoucherCode.code` (unique, max_length=50)
  - name         optional — `VoucherCode.name`, falls back to a sensible
                              default if unmapped (the model requires it)
  - description  optional — `VoucherCode.description`
  - external_id  optional — `VoucherCode.external_id` (e.g. One 15
                              Marina's member identifier)

Everything else on `VoucherCode` (discount_type, discount_value, dates,
usage limits, the four restriction booleans, etc.) comes from the
admin's batch-level form. We never honour those fields per-row, which
keeps the column-mapping UI compact and the import semantics easy to
reason about.

Duplicate handling: callers pick one of
  - "skip"      keep existing rows untouched, drop file rows whose
                code already exists
  - "overwrite" update existing rows' settings with the batch settings
                (codes, current_uses, created_at all preserved)
  - "fail"      abort the whole import with a `ValidationError` if any
                file code is already in use

Limits: hard caps of 5 MB upload size and 5 000 rows per batch. Larger
imports should be queued through the migration Celery infrastructure
in a follow-up — `MigrationJob` is already linked from `VoucherCode`
for exactly that upgrade path.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from typing import Any, Iterable

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _

from migration.models import MigrationJob
from vouchers.models import VoucherCode

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
MAX_ROWS = 5_000
ALLOWED_EXTENSIONS = (".csv", ".xlsx")

#: Mappable fields the importer accepts as targets, plus how the column
#: auto-detector normalises common header spellings to each target.
#: Only `code` is required; the others fall back to defaults if unmapped.
MAPPABLE_FIELDS: dict[str, dict[str, Any]] = {
    "code": {
        "label": _("Voucher code"),
        "required": True,
        "max_length": 50,
        "aliases": (
            "code", "voucher_code", "coupon_code", "voucher", "coupon",
            "promo_code", "discount_code", "vouchercode", "couponcode",
        ),
    },
    "name": {
        "label": _("Internal name"),
        "required": False,
        "max_length": 200,
        "aliases": (
            "name", "voucher_name", "title", "internal_name",
            "description_short", "campaign",
        ),
    },
    "description": {
        "label": _("Customer-facing description"),
        "required": False,
        "max_length": None,
        "aliases": ("description", "details", "long_description", "note"),
    },
    "external_id": {
        "label": _("External ID"),
        "required": False,
        "max_length": 100,
        "aliases": (
            "external_id", "externalid", "source_id", "member_id",
            "customer_id", "user_id", "external_ref", "reference",
        ),
    },
}

DUPLICATE_STRATEGIES = ("skip", "overwrite", "fail")

#: Column order used for both the export and the suggested re-import
#: shape. Keeping these aligned is what makes export → edit → re-import
#: a stable round-trip; see `tests/unit/test_voucher_import_export.py`.
EXPORT_COLUMNS: tuple[str, ...] = (
    "code", "name", "description", "external_id",
    "discount_type", "discount_value", "max_discount_amount",
    "application_scope", "start_date", "end_date", "days_valid",
    "max_uses_total", "max_uses_per_customer", "current_uses",
    "min_order_value", "exclude_sale_items",
    "cannot_combine_with_other_vouchers", "cannot_combine_with_sale_items",
    "first_time_customers_only", "is_active", "created_at",
)


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclass
class ParsedFile:
    """What `parse_file` returns to the admin's preview page."""

    headers: list[str]
    rows: list[dict[str, str]]
    row_count: int

    def first(self, n: int) -> list[dict[str, str]]:
        return self.rows[:n]


@dataclass
class InvalidRow:
    """A file row we couldn't accept, surfaced to the admin so they can
    fix the source file and retry."""

    row_number: int  # 1-based, header is row 1
    code: str
    reason: str


@dataclass
class ImportPreview:
    """What the admin sees on the preview page before committing."""

    parsed: ParsedFile
    mapping: dict[str, str]
    valid_count: int
    invalid_rows: list[InvalidRow]
    existing_codes: set[str]
    sample_duplicates: list[str]

    @property
    def duplicate_count(self) -> int:
        return len(self.existing_codes)

    @property
    def new_count(self) -> int:
        return self.valid_count - self.duplicate_count


@dataclass
class ImportResult:
    """Outcome of a confirmed import, persisted to the result page."""

    imported: int = 0
    updated: int = 0
    skipped_duplicate: int = 0
    skipped_invalid: int = 0
    migration_job_id: str | None = None
    errors: list[str] = field(default_factory=list)

    @property
    def total_processed(self) -> int:
        return self.imported + self.updated + self.skipped_duplicate + self.skipped_invalid


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _normalise_header(header: str | None) -> str:
    """Lowercase and squash punctuation/whitespace for header matching."""
    if not header:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", str(header).strip().lower()).strip("_")


def parse_file(uploaded_file) -> ParsedFile:
    """Read a CSV or XLSX upload into a normalised (headers, rows) shape.

    `uploaded_file` is a Django `UploadedFile` from a multipart form. Both
    branches strip whitespace, drop entirely-blank rows, and stringify
    every cell so the column mapping that runs next can treat the value
    pool uniformly (the model layer does the eventual type coercion).
    """
    name = (getattr(uploaded_file, "name", "") or "").lower()
    if not name.endswith(ALLOWED_EXTENSIONS):
        raise ValidationError(
            _("Unsupported file type. Upload a .csv or .xlsx file.")
        )

    size = getattr(uploaded_file, "size", None)
    if size is not None and size > MAX_UPLOAD_BYTES:
        raise ValidationError(
            _("File is too large. Maximum upload size is 5 MB.")
        )

    if name.endswith(".csv"):
        headers, rows = _parse_csv(uploaded_file)
    else:
        headers, rows = _parse_xlsx(uploaded_file)

    if len(rows) > MAX_ROWS:
        raise ValidationError(
            _("Too many rows ({count}). Maximum is {limit} per import.").format(
                count=len(rows), limit=MAX_ROWS,
            )
        )

    return ParsedFile(headers=headers, rows=rows, row_count=len(rows))


def _parse_csv(uploaded_file) -> tuple[list[str], list[dict[str, str]]]:
    raw = uploaded_file.read()
    # Reset the pointer in case the caller (or our preview-cache loop)
    # wants to read it again later.
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    text = raw.decode("utf-8-sig", errors="replace") if isinstance(raw, bytes) else raw

    # Sniff the delimiter, but fall back to comma if the sniffer chokes
    # (single-column files, etc.).
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows_iter = iter(reader)
    try:
        headers = [str(h).strip() for h in next(rows_iter)]
    except StopIteration:
        return [], []

    out: list[dict[str, str]] = []
    for row in rows_iter:
        cells = [str(v).strip() if v is not None else "" for v in row]
        if not any(cells):
            continue
        # Pad/truncate to header width so dict zipping is total.
        cells = (cells + [""] * len(headers))[: len(headers)]
        out.append(dict(zip(headers, cells)))
    return headers, out


def _parse_xlsx(uploaded_file) -> tuple[list[str], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    ws = wb.active
    if ws is None:
        return [], []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    out: list[dict[str, str]] = []
    for row in rows_iter:
        cells = [_cell_to_str(v) for v in row]
        if not any(c for c in cells):
            continue
        cells = (cells + [""] * len(headers))[: len(headers)]
        out.append(dict(zip(headers, cells)))
    return headers, out


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


# ---------------------------------------------------------------------------
# Mapping
# ---------------------------------------------------------------------------

def auto_detect_mapping(headers: Iterable[str]) -> dict[str, str]:
    """Suggest a {target_field: source_header} mapping from headers.

    Compares each header's normalised form against each field's alias set
    and returns the first matches. Admins can override every choice on
    the preview page; this is only a starting point so the most common
    files (with a `Code` column) don't need any clicks.
    """
    suggestion: dict[str, str] = {}
    normalised = [(_normalise_header(h), h) for h in headers]
    for target, spec in MAPPABLE_FIELDS.items():
        alias_set = set(spec["aliases"])
        for norm, original in normalised:
            if norm in alias_set:
                suggestion[target] = original
                break
    return suggestion


def validate_mapping(mapping: dict[str, str], headers: Iterable[str]) -> dict[str, str]:
    """Drop unknown targets, drop empty values, ensure `code` is set."""
    headers_set = set(headers)
    cleaned: dict[str, str] = {}
    for target, source in mapping.items():
        if target not in MAPPABLE_FIELDS:
            continue
        if not source:
            continue
        if source not in headers_set:
            raise ValidationError(
                _("Column '{source}' not found in the uploaded file.").format(
                    source=source
                )
            )
        cleaned[target] = source

    if "code" not in cleaned:
        raise ValidationError(_("You must map a column to the voucher code."))
    return cleaned


# ---------------------------------------------------------------------------
# Validation + preview
# ---------------------------------------------------------------------------

def _row_to_target_values(
    row: dict[str, str], mapping: dict[str, str]
) -> dict[str, str]:
    """Translate one raw file row into a `{target_field: value}` dict."""
    return {target: (row.get(source) or "").strip() for target, source in mapping.items()}


def partition_rows(
    rows: list[dict[str, str]], mapping: dict[str, str]
) -> tuple[list[dict[str, str]], list[InvalidRow]]:
    """Split rows into the ones we can accept and the ones we have to
    reject (with a human-readable reason)."""
    valid: list[dict[str, str]] = []
    invalid: list[InvalidRow] = []
    seen_in_batch: set[str] = set()

    for idx, row in enumerate(rows, start=2):  # row 1 is the header
        targets = _row_to_target_values(row, mapping)
        code = targets.get("code", "")
        if not code:
            invalid.append(InvalidRow(idx, code, str(_("Code is empty"))))
            continue
        if len(code) > MAPPABLE_FIELDS["code"]["max_length"]:
            invalid.append(InvalidRow(idx, code, str(_("Code exceeds 50 characters"))))
            continue

        # Per-field length checks for the optional mappable targets.
        ok = True
        for target, max_len in (("name", 200), ("external_id", 100)):
            val = targets.get(target, "")
            if val and max_len and len(val) > max_len:
                invalid.append(InvalidRow(
                    idx, code,
                    str(_("{target} exceeds {max_len} characters")).format(
                        target=target, max_len=max_len,
                    ),
                ))
                ok = False
                break
        if not ok:
            continue

        if code in seen_in_batch:
            invalid.append(InvalidRow(idx, code, str(_("Code appears more than once in the file"))))
            continue
        seen_in_batch.add(code)
        valid.append(targets)

    return valid, invalid


def find_existing_codes(codes: Iterable[str]) -> set[str]:
    """Return the subset of `codes` that already exist in the DB."""
    codes_list = list(codes)
    if not codes_list:
        return set()
    return set(
        VoucherCode.objects
        .filter(code__in=codes_list)
        .values_list("code", flat=True)
    )


def build_preview(parsed: ParsedFile, mapping: dict[str, str]) -> ImportPreview:
    valid, invalid = partition_rows(parsed.rows, mapping)
    existing = find_existing_codes(row["code"] for row in valid)
    sample = sorted(existing)[:10]
    return ImportPreview(
        parsed=parsed,
        mapping=mapping,
        valid_count=len(valid),
        invalid_rows=invalid,
        existing_codes=existing,
        sample_duplicates=sample,
    )


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def _normalise_batch_settings(settings: dict[str, Any]) -> dict[str, Any]:
    """Strip blank values so the model defaults / `null=True` columns
    don't get overwritten with empty strings."""
    cleaned: dict[str, Any] = {}
    for key, value in settings.items():
        if value in ("", None):
            continue
        cleaned[key] = value
    return cleaned


@transaction.atomic
def import_batch(
    parsed: ParsedFile,
    mapping: dict[str, str],
    batch_settings: dict[str, Any],
    dup_strategy: str,
    user,
    *,
    job: MigrationJob | None = None,
) -> ImportResult:
    """Apply a validated batch. Returns counts for the result page."""
    if dup_strategy not in DUPLICATE_STRATEGIES:
        raise ValidationError(
            _("Unknown duplicate strategy: {strategy}").format(strategy=dup_strategy)
        )

    mapping = validate_mapping(mapping, parsed.headers)
    valid, invalid = partition_rows(parsed.rows, mapping)
    existing = find_existing_codes(row["code"] for row in valid)
    settings = _normalise_batch_settings(batch_settings)

    if dup_strategy == "fail" and existing:
        raise ValidationError(
            _("Import aborted: {count} codes already exist in the catalogue.").format(
                count=len(existing)
            )
        )

    if job is None:
        job = MigrationJob.objects.create(
            created_by=user,
            platform="csv",
            method="csv",
            import_products=False,
            import_categories=False,
            import_customers=False,
            import_orders=False,
            import_reviews=False,
            import_coupons=True,
            status="running",
            current_step="voucher_import",
            coupons_total=len(valid),
        )

    result = ImportResult(
        skipped_invalid=len(invalid),
        migration_job_id=str(job.id),
    )

    to_create: list[VoucherCode] = []
    codes_to_update: list[str] = []

    for row in valid:
        code = row["code"]
        if code in existing:
            if dup_strategy == "skip":
                result.skipped_duplicate += 1
                continue
            if dup_strategy == "overwrite":
                codes_to_update.append(code)
            continue
        to_create.append(VoucherCode(
            code=code,
            name=row.get("name") or f"Imported voucher {code}",
            description=row.get("description", ""),
            external_id=row.get("external_id", ""),
            created_by=user,
            migration_job=job,
            **settings,
        ))

    if to_create:
        VoucherCode.objects.bulk_create(to_create, batch_size=500)
        result.imported = len(to_create)

    if codes_to_update:
        # Build a per-code update payload: optional column values from the
        # row override the batch settings (so an admin who maps a `name`
        # column can rename existing rows in place). Everything that
        # touches identity / lifecycle (code, current_uses, created_at,
        # created_by) is intentionally excluded so re-imports don't break
        # audit trails.
        for code in codes_to_update:
            payload = dict(settings)
            mapped_row = next((row for row in valid if row["code"] == code), {})
            for opt in ("name", "description", "external_id"):
                if opt in mapped_row and mapped_row[opt]:
                    payload[opt] = mapped_row[opt]
            payload["migration_job_id"] = job.id
            VoucherCode.objects.filter(code=code).update(**payload)
        result.updated = len(codes_to_update)

    job.coupons_imported = result.imported
    job.coupons_skipped = result.skipped_duplicate + result.skipped_invalid
    job.status = "completed"
    job.progress_percent = 100
    job.save(update_fields=[
        "coupons_imported", "coupons_skipped", "status", "progress_percent",
    ])

    logger.info(
        "Voucher import complete: job=%s imported=%d updated=%d skipped_dup=%d skipped_invalid=%d",
        job.id, result.imported, result.updated,
        result.skipped_duplicate, result.skipped_invalid,
    )
    return result


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def _voucher_row(voucher: VoucherCode) -> list[Any]:
    """Project a `VoucherCode` onto `EXPORT_COLUMNS`."""
    def money_amount(field):
        value = getattr(voucher, field)
        return str(value.amount) if value and hasattr(value, "amount") else ""

    return [
        voucher.code,
        voucher.name,
        voucher.description,
        voucher.external_id or "",
        voucher.discount_type,
        str(voucher.discount_value) if voucher.discount_value is not None else "",
        money_amount("max_discount_amount"),
        voucher.application_scope,
        voucher.start_date.isoformat() if voucher.start_date else "",
        voucher.end_date.isoformat() if voucher.end_date else "",
        voucher.days_valid if voucher.days_valid is not None else "",
        voucher.max_uses_total if voucher.max_uses_total is not None else "",
        voucher.max_uses_per_customer if voucher.max_uses_per_customer is not None else "",
        voucher.current_uses,
        money_amount("min_order_value"),
        voucher.exclude_sale_items,
        voucher.cannot_combine_with_other_vouchers,
        voucher.cannot_combine_with_sale_items,
        voucher.first_time_customers_only,
        voucher.is_active,
        voucher.created_at.isoformat() if voucher.created_at else "",
    ]


def export_queryset(queryset, fmt: str = "csv", filename: str = "vouchers") -> HttpResponse:
    """Stream a CSV or XLSX response of `queryset` projected onto
    `EXPORT_COLUMNS`. The column order matches what the importer accepts
    so round-tripping (export → edit → re-import) just works."""
    if fmt == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(EXPORT_COLUMNS)
        for voucher in queryset.iterator(chunk_size=500):
            writer.writerow(_voucher_row(voucher))
        return response

    if fmt == "xlsx":
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title="Vouchers")
        ws.append(list(EXPORT_COLUMNS))
        for voucher in queryset.iterator(chunk_size=500):
            ws.append(_voucher_row(voucher))
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response

    raise ValidationError(_("Unknown export format: {fmt}").format(fmt=fmt))
